#!/usr/bin/env python3
"""Экспорт в xlsx: реализованные фичи стека + доработки (утечки JSON, TTS, видео STT)."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "Реализованные_фичи.xlsx"

# Из scripts/build_feature_template_xlsx.py — только «Сделано» / «Сделано*»
TZ_DONE: list[tuple[int, str, str, str, str]] = [
    (1, "Текстовой чат", "обязательная", "Сделано", "Open WebUI + gpthub-gateway POST /v1/chat/completions → MWS."),
    (
        2,
        "Голосовой чат",
        "обязательная",
        "Сделано",
        "Запись/звонок → STT → тот же чат → опционально TTS ответа.",
    ),
    (
        3,
        "Генерация изображений в чате",
        "обязательная",
        "Сделано*",
        "Перехват gena: stream_image_markdown → /images/generations.",
    ),
    (
        4,
        "Аудиофайлы + автоматический ASR",
        "обязательная",
        "Сделано",
        "Вложения → /v1/audio/transcriptions (прокси на MWS/Whisper).",
    ),
    (5, "Изображения (VLM)", "обязательная", "Сделано", "image_url → vision-модель из каталога."),
    (6, "Файлы и ответы по содержимому", "обязательная", "Сделано", "RAG, чанки в контекст."),
    (7, "Поиск в интернете", "обязательная", "Сделано", "DuckDuckGo + сниппеты в контекст."),
    (8, "Веб-парсинг по ссылке", "обязательная", "Сделано", "URL → trafilatura/fetch текста."),
    (9, "Долгосрочная память", "обязательная", "Сделано", "SQLite + эмбеддинги, Chroma опционально."),
    (10, "Автовыбор модели под задачу", "обязательная", "Сделано", "gpthub-auto → pick_route_gena + перехваты."),
    (11, "Ручной выбор модели", "обязательная", "Сделано", "apply_manual_route по id из /v1/models."),
    (12, "Markdown и форматированный код", "обязательная", "Сделано", "Рендер Open WebUI."),
    (
        13,
        "Deep Research",
        "дополнительная",
        "Сделано",
        "stream_deep_research: поиск + отчёт LLM.",
    ),
    (
        14,
        "Генерация презентаций",
        "дополнительная",
        "Сделано",
        "stream_presentation_pptx: python-pptx, PDF при необходимости.",
    ),
    (
        15,
        "TTS + прочее (шлюз)",
        "дополнительная",
        "Сделано",
        "Прокси /v1/audio/speech; music_demo (MP3) по триггерам.",
    ),
]

# Доработки из репозитория (интеграционные исправления)
RECENT: list[tuple[str, str, str, str]] = [
    (
        "gpthub-gateway",
        "Подавление утечки JSON Open WebUI в ответ чата",
        "Сделано",
        "Если модель вернула только {\"queries\": [...]} или {\"follow_ups\": [...]} — повторный запрос к MWS с инструкцией отвечать текстом; stream + non-stream; тесты try_parse_*.",
    ),
    (
        "open-webui-src",
        "TTS: прокси /openai/audio/speech с кастомным API URL",
        "Сделано",
        "Раньше искался только https://api.openai.com/v1; для gpthub-gateway используется индекс 0 из OPENAI_API_BASE_URLS.",
    ),
    (
        "open-webui-src",
        "Обработка вложений video/mp4 для STT",
        "Сделано",
        "strict_match_mime_type: по умолчанию audio/* и video/* (раньше только video/webm).",
    ),
]


def _style_header(ws, row: int, ncols: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main() -> None:
    wb = Workbook()

    # --- Лист 1: ТЗ, реализовано ---
    ws1 = wb.active
    ws1.title = "ТЗ реализовано"
    h1 = ("№", "Функционал", "Тип", "Статус", "Комментарий")
    for col, h in enumerate(h1, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header(ws1, 1, len(h1))
    for i, row in enumerate(TZ_DONE, start=2):
        for j, val in enumerate(row, start=1):
            c = ws1.cell(row=i, column=j, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 44
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 58
    ws1.freeze_panes = "A2"

    # --- Лист 2: доработки ---
    ws2 = wb.create_sheet("Доработки интеграции")
    h2 = ("Компонент", "Фича / исправление", "Статус", "Описание")
    for col, h in enumerate(h2, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(h2))
    for i, row in enumerate(RECENT, start=2):
        for j, val in enumerate(row, start=1):
            c = ws2.cell(row=i, column=j, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 48
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 72
    ws2.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Written: {OUT}")


if __name__ == "__main__":
    main()
