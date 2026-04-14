#!/usr/bin/env python3
"""Сборка GPTHub шаблон фич.xlsx — таблица статусов и дорожная карта."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "GPTHub шаблон фич (1).xlsx"

# №, Функционал, Тип, Статус, Комментарий (как в проекте / ограничения)
ROWS: list[tuple[int, str, str, str, str]] = [
    # —— обязательные из ТЗ ——
    (1, "Текстовой чат", "обязательная", "Сделано", "Open WebUI + gpthub-gateway POST /v1/chat/completions → MWS."),
    (
        2,
        "Голосовой чат",
        "обязательная",
        "Сделано",
        "Запись/звонок → STT (Whisper/OpenAI-совместимо) → тот же чат → при желании TTS ответа. Не отдельный голосовой диалог с моделью «как телефон».",
    ),
    (
        3,
        "Генерация изображений в чате",
        "обязательная",
        "Сделано*",
        "Перехват gena: stream_image_markdown → /images/generations, промпт через LLM. *Составление сцены с предыдущей картинкой — через текстовый промпт (референс в контексте), не отдельный img2img в шлюзе.",
    ),
    (
        4,
        "Аудиофайлы + автоматический ASR",
        "обязательная",
        "Сделано",
        "Вложения → /v1/audio/transcriptions (прокси на MWS/Whisper по настройке compose).",
    ),
    (
        5,
        "Изображения (VLM)",
        "обязательная",
        "Сделано",
        "image_url в сообщении → роутер выбирает vision-модель из каталога.",
    ),
    (
        6,
        "Файлы и ответы по содержимому",
        "обязательная",
        "Сделано",
        "RAG: эмбеддинги, SQLite/Chroma по настройке, чанки в контекст (см. memory_context, RAG в шлюзе).",
    ),
    (
        7,
        "Поиск в интернете",
        "обязательная",
        "Сделано",
        "DuckDuckGo + встраивание сниппетов в контекст (web_tools / роутер).",
    ),
    (
        8,
        "Веб-парсинг по ссылке из сообщения",
        "обязательная",
        "Сделано",
        "Извлечение URL, trafilatura/fetch текста на страницу.",
    ),
    (
        9,
        "Долгосрочная память",
        "обязательная",
        "Сделано",
        "SQLite + эмбеддинги, опционально Chroma; явные «запомни» и фоновый digest — по реализации шлюза.",
    ),
    (
        10,
        "Автовыбор модели под задачу",
        "обязательная",
        "Сделано",
        "gpthub-auto → pick_route_gena + перехваты (презентация, картинка, deep research и т.д.).",
    ),
    (
        11,
        "Ручной выбор модели",
        "обязательная",
        "Сделано",
        "Конкретный id из /v1/models, apply_manual_route.",
    ),
    (
        12,
        "Markdown и форматированный код",
        "обязательная",
        "Сделано",
        "Рендер Open WebUI (код-блоки, таблицы, картинки).",
    ),
    # —— дополнительные из ТЗ ——
    (
        13,
        "Deep Research / research-режим",
        "дополнительная",
        "Сделано",
        "stream_deep_research: поиск + страницы + отчёт LLM (gena_long_doc_model).",
    ),
    (
        14,
        "Генерация презентаций",
        "дополнительная",
        "Сделано",
        "stream_presentation_pptx: JSON слайдов → python-pptx, PDF при необходимости; стиль из UI/промпта.",
    ),
    (
        15,
        "Другой доп. функционал (из репозитория)",
        "дополнительная",
        "Сделано",
        "TTS: прокси /v1/audio/speech; демо-генерация MP3-мелодии (music_demo) по триггерам — см. FEATURES_CHECKLIST.md.",
    ),
    # —— предлагаемые обязательные к развитию (сейчас частично или на стороне инфраструктуры) ——
    (
        16,
        "Корпоративный вход (SSO: OIDC/SAML) и роли",
        "обязательная (к развитию)",
        "Частично",
        "В Open WebUI есть механизмы авторизации; единая политика с IdP заказчика и роли admin/user — по внедрению.",
    ),
    (
        17,
        "Аудит: кто и какие запросы к моделям (логирование)",
        "обязательная (к развитию)",
        "Частично",
        "Логи шлюза/Traefik; полноценный аудит с хранением и отчётами — при необходимости отдельный сервис.",
    ),
    (
        18,
        "Квоты и учёт потребления токенов/API",
        "обязательная (к развитию)",
        "Не сделано",
        "В шлюзе нет встроенного биллинга; возможна интеграция на уровне API-ключей MWS/прокси.",
    ),
    (
        19,
        "Наблюдаемость: health, метрики, алерты",
        "обязательная (к развитию)",
        "Частично",
        "/health у сервисов; полный стек мониторинга (Prometheus/Grafana) — по инфраструктуре.",
    ),
    # —— предлагаемые дополнительные фичи ——
    (
        20,
        "Настоящий img2img / inpainting по загруженной картинке",
        "дополнительная (к развитию)",
        "Не сделано",
        "Сейчас правки через текстовый промпт и ссылки; нужна поддержка API с image+mask при наличии у провайдера.",
    ),
    (
        21,
        "Инструменты агента: стабильный вызов внешних API (tools/functions)",
        "дополнительная (к развитию)",
        "Частично",
        "Зависит от моделей MWS и настроек Open WebUI; расширяемый сценарий.",
    ),
    (
        22,
        "Импорт корпоративных источников (Confluence, SharePoint, Notion)",
        "дополнительная (к развитию)",
        "Не сделано",
        "RAG сейчас от загружаемых файлов/чата; коннекторы — отдельный проект.",
    ),
    (
        23,
        "Экспорт чата / отчёт для compliance (PDF, архив)",
        "дополнительная (к развитию)",
        "Частично",
        "Экспорт есть в Open WebUI; единый формат под аудит — по требованиям.",
    ),
    (
        24,
        "Code interpreter (выполнение кода в изолированной среде)",
        "дополнительная (к развитию)",
        "Частично",
        "Зависит от включения в Open WebUI и бэкенда; не ядро gpthub-gateway.",
    ),
    (
        25,
        "Мультиязычный UI и локализация под бренд заказчика",
        "дополнительная (к развитию)",
        "Частично",
        "Open WebUI мультиязычен; кастомизация бренда — темы/статика.",
    ),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Фичи GPTHub"

    headers = ("№", "Функционал", "Тип", "Статус", "Комментарий")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, row in enumerate(ROWS, start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 62

    ws.freeze_panes = "A2"
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Written: {OUT}")


if __name__ == "__main__":
    main()
